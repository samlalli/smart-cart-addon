"""
exports.py — Excel, PDF and plain text export for Smart Cart v1.1.0
"""
import io
import os
from datetime import datetime


def format_price(price):
    if price is None:
        return "—"
    if price >= 1000:
        return f"${price:,.2f}"
    return f"${price:.2f}"


def export_excel(list_data: dict, recipes_data: dict, history_data: dict, store_lists: dict = None) -> bytes:
    """Generate Excel file with multiple tabs."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Tab 1: Shopping List ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Shopping List"

    headers1 = ["Item", "Qty", "Unit", "Details", "Category", "Notes",
                 "Included This Week", "Cupboard Item", "Source",
                 "Purchase Count", "Last Purchased"]
    header_fill = PatternFill("solid", fgColor="1a2010")
    header_font = Font(bold=True, color="c8f060")

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(list_data.get("items", []), 2):
        ws1.cell(row=row, column=1, value=item.get("item", ""))
        ws1.cell(row=row, column=2, value=item.get("qty", 1))
        ws1.cell(row=row, column=3, value=item.get("unit", ""))
        ws1.cell(row=row, column=4, value=item.get("details", ""))
        ws1.cell(row=row, column=5, value=item.get("category", ""))
        ws1.cell(row=row, column=6, value=item.get("notes", ""))
        ws1.cell(row=row, column=7, value="Yes" if item.get("included") else "No")
        ws1.cell(row=row, column=8, value="Yes" if item.get("is_cupboard") else "No")
        sources = item.get("sources", [])
        tags = item.get("recipe_tags", [])
        source_label = "Recipe" if "recipe" in sources and tags else "Manual"
        ws1.cell(row=row, column=9, value=source_label)
        ws1.cell(row=row, column=10, value=item.get("purchase_count", 0))
        last = item.get("last_purchased", "")
        if last:
            try:
                last = datetime.fromisoformat(last).strftime("%d/%m/%Y")
            except Exception:
                pass
        ws1.cell(row=row, column=11, value=last)

    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].auto_size = True
        ws1.column_dimensions[get_column_letter(col)].width = max(12, 20)

    # ── Tab 2: Recipes ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Recipes")
    headers2 = ["Recipe Name", "Base Servings", "Cook Count", "Last Cooked",
                 "Source URL", "Ingredient", "Qty", "Unit", "Details",
                 "Is Pantry/Minor", "Category", "Notes"]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for recipe in recipes_data.get("recipes", []):
        last = recipe.get("last_cooked", "")
        if last:
            try:
                last = datetime.fromisoformat(last).strftime("%d/%m/%Y")
            except Exception:
                pass
        ingredients = recipe.get("ingredients", [])
        if not ingredients:
            ws2.cell(row=row, column=1, value=recipe.get("name", ""))
            ws2.cell(row=row, column=2, value=recipe.get("base_servings", 4))
            ws2.cell(row=row, column=3, value=recipe.get("cook_count", 0))
            ws2.cell(row=row, column=4, value=last)
            ws2.cell(row=row, column=5, value=recipe.get("source_url", ""))
            row += 1
        else:
            for i, ing in enumerate(ingredients):
                ws2.cell(row=row, column=1, value=recipe.get("name", "") if i == 0 else "")
                ws2.cell(row=row, column=2, value=recipe.get("base_servings", 4) if i == 0 else "")
                ws2.cell(row=row, column=3, value=recipe.get("cook_count", 0) if i == 0 else "")
                ws2.cell(row=row, column=4, value=last if i == 0 else "")
                ws2.cell(row=row, column=5, value=recipe.get("source_url", "") if i == 0 else "")
                ws2.cell(row=row, column=6, value=ing.get("item", ""))
                ws2.cell(row=row, column=7, value=ing.get("qty", 1))
                ws2.cell(row=row, column=8, value=ing.get("unit", ""))
                ws2.cell(row=row, column=9, value=ing.get("details", ""))
                ws2.cell(row=row, column=10, value="Yes" if (ing.get("is_pantry") or ing.get("is_small_qty")) else "No")
                ws2.cell(row=row, column=11, value=ing.get("category", ""))
                ws2.cell(row=row, column=12, value=recipe.get("notes", "") if i == 0 else "")
                row += 1

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ── Tab 3: Shop History ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Shop History")
    headers3 = ["Date", "Stores", "Items", "Subtotal", "Delivery",
                 "Total", "Saved", "Rewards Plus Saved", "Rewards Plus Applied"]

    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, shop in enumerate(history_data.get("shops", []), 2):
        date = shop.get("date", "")
        if date:
            try:
                date = datetime.fromisoformat(date).strftime("%d/%m/%Y %H:%M")
            except Exception:
                pass
        ws3.cell(row=row, column=1, value=date)
        ws3.cell(row=row, column=2, value=", ".join(s.capitalize() for s in shop.get("stores", [])))
        ws3.cell(row=row, column=3, value=shop.get("item_count", ""))
        ws3.cell(row=row, column=4, value=shop.get("subtotal", 0))
        ws3.cell(row=row, column=5, value=shop.get("delivery", 0))
        ws3.cell(row=row, column=6, value=shop.get("total", 0))
        ws3.cell(row=row, column=7, value=shop.get("saved", 0))
        ws3.cell(row=row, column=8, value=shop.get("rewards_plus_saved", 0))
        ws3.cell(row=row, column=9, value="Yes" if shop.get("rewards_plus_applied") else "No")

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # ── Tab 4+: In-store lists ────────────────────────────────────────────────
    if store_lists:
        store_colors = {
            "woolworths": ("0d3318", "4dde7e"),
            "coles": ("3d0a09", "f07070"),
            "aldi": ("0d1e3d", "6090e8")
        }
        for store, items in store_lists.items():
            if not items:
                continue
            ws = wb.create_sheet(f"{store.capitalize()} List")
            sc = store_colors.get(store, ("1a1a1a", "ffffff"))
            sf = PatternFill("solid", fgColor=sc[0])
            sfont = Font(bold=True, color=sc[1])

            headers_s = ["Item", "Matched Product", "Qty", "Unit", "Details",
                         "Category", "Aisle", "Unit Price", "Total", "On Special", "Estimated"]
            for col, h in enumerate(headers_s, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = sfont
                cell.fill = sf
                cell.alignment = Alignment(horizontal="center")

            # Sort by aisle then category
            sorted_items = sorted(items, key=lambda x: (x.get("aisle") or x.get("category") or "zzz"))
            grand_total = 0

            for row, item in enumerate(sorted_items, 2):
                # price is already the line total; buy_qty is the practical purchase quantity
                line_total = item.get("price")
                unit_price = item.get("unit_price", line_total)
                buy_qty = item.get("buy_qty") or str(item.get("packs", 1))
                if line_total:
                    grand_total += line_total
                ws.cell(row=row, column=1, value=item.get("item", ""))
                ws.cell(row=row, column=2, value=item.get("matched_name", ""))
                ws.cell(row=row, column=3, value=buy_qty)
                ws.cell(row=row, column=4, value=item.get("recipe_unit", item.get("unit", "")))
                ws.cell(row=row, column=5, value=item.get("details", ""))
                ws.cell(row=row, column=6, value=item.get("category", ""))
                ws.cell(row=row, column=7, value=item.get("aisle", ""))
                ws.cell(row=row, column=8, value=unit_price)
                ws.cell(row=row, column=9, value=line_total)
                ws.cell(row=row, column=10, value="Yes" if item.get("on_special") else "")
                ws.cell(row=row, column=11, value="Est." if item.get("estimated") else "")

            # Total row
            total_row = len(sorted_items) + 2
            ws.cell(row=total_row, column=8, value="TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=round(grand_total, 2)).font = Font(bold=True)

            for col in range(1, len(headers_s) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_pdf(list_data: dict, settings: dict = None) -> bytes:
    """Generate A4 portrait PDF of this week's included items."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=20*mm, rightMargin=20*mm,
                             topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    accent = colors.HexColor("#c8f060")
    dark = colors.HexColor("#1a1a1a")
    muted = colors.HexColor("#888888")

    title_style = ParagraphStyle("Title", fontSize=22, fontName="Helvetica-Bold",
                                  textColor=dark, spaceAfter=4)
    date_style = ParagraphStyle("Date", fontSize=10, fontName="Helvetica",
                                 textColor=muted, spaceAfter=16)
    cat_style = ParagraphStyle("Cat", fontSize=12, fontName="Helvetica-Bold",
                                textColor=dark, spaceBefore=12, spaceAfter=6)
    item_style = ParagraphStyle("Item", fontSize=10, fontName="Helvetica",
                                 textColor=dark, leftIndent=10)
    recipe_style = ParagraphStyle("Recipe", fontSize=8, fontName="Helvetica-Oblique",
                                   textColor=muted, leftIndent=20)

    # Get included items
    items = [i for i in list_data.get("items", []) if i.get("included", True)]

    # Group by category
    categories = {}
    for item in items:
        cat = item.get("category") or "Other"
        categories.setdefault(cat, []).append(item)

    story = []
    story.append(Paragraph("🛒 Smart Cart", title_style))
    story.append(Paragraph(f"Shopping List — {datetime.now().strftime('%d %B %Y')}", date_style))
    story.append(HRFlowable(width="100%", thickness=2, color=accent, spaceAfter=12))

    total_items = 0
    for cat, cat_items in sorted(categories.items()):
        story.append(Paragraph(cat.upper(), cat_style))
        for item in cat_items:
            qty = item.get("qty", 1)
            unit = item.get("unit", "")
            details = item.get("details", "")
            name = item.get("item", "")
            notes = item.get("notes", "")

            qty_str = f"{qty:g}" if qty != int(qty) else str(int(qty))
            unit_str = f" {unit}" if unit else ""
            details_str = f" — {details}" if details else ""
            notes_str = f" ({notes})" if notes else ""

            story.append(Paragraph(f"• {name}{details_str} × {qty_str}{unit_str}{notes_str}", item_style))

            tags = item.get("recipe_tags", [])
            if tags:
                story.append(Paragraph(f"For recipe", recipe_style))

            total_items += 1

    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", thickness=1, color=muted, spaceAfter=8))
    story.append(Paragraph(f"Total items: {total_items}", date_style))

    doc.build(story)
    buf.seek(0)
    return buf.read()


def export_plain_text(list_data: dict, store_lists: dict = None, mode: str = "list") -> str:
    """
    Generate plain text for sharing.
    mode: 'list' = this week's list, 'store' = in-store list by store
    """
    lines = []
    date_str = datetime.now().strftime("%d %B %Y")

    if mode == "list" or not store_lists:
        lines.append(f"🛒 Smart Cart — {date_str}")
        lines.append("")

        items = [i for i in list_data.get("items", []) if i.get("included", True)]
        categories = {}
        for item in items:
            cat = item.get("category") or "Other"
            categories.setdefault(cat, []).append(item)

        for cat, cat_items in sorted(categories.items()):
            lines.append(cat.upper())
            for item in cat_items:
                qty = item.get("qty", 1)
                unit = item.get("unit", "")
                details = item.get("details", "")
                name = item.get("item", "")
                notes = item.get("notes", "")

                qty_str = f"{qty:g}" if qty != int(qty) else str(int(qty))
                unit_str = f" {unit}" if unit else ""
                details_str = f" ({details})" if details else ""
                notes_str = f" — {notes}" if notes else ""

                lines.append(f"• {name}{details_str} × {qty_str}{unit_str}{notes_str}")
            lines.append("")

    elif mode == "store" and store_lists:
        store_emoji = {"woolworths": "🟢", "coles": "🔴", "aldi": "🔵"}
        for store, items in store_lists.items():
            if not items:
                continue
            emoji = store_emoji.get(store, "🛒")
            lines.append(f"{emoji} {store.upper()} — {date_str}")
            lines.append("")

            # Group by aisle or category
            groups = {}
            for item in items:
                group = item.get("aisle") or item.get("category") or "Other"
                groups.setdefault(group, []).append(item)

            grand_total = 0
            for group, group_items in sorted(groups.items()):
                lines.append(f"  {group}")
                for item in group_items:
                    line_total = item.get("price")  # already the line total
                    if line_total:
                        grand_total += line_total
                    buy_qty = item.get("buy_qty") or str(item.get("packs", 1))
                    price_str = format_price(line_total) if line_total else "—"
                    est_str = " (est.)" if item.get("estimated") else ""
                    special_str = " ⭐ ON SPECIAL" if item.get("on_special") else ""

                    lines.append(f"  • {item.get('item', '')} — {buy_qty} — {price_str}{est_str}{special_str}")
                lines.append("")

            lines.append(f"  TOTAL: {format_price(grand_total)}")
            lines.append("")

    return "\n".join(lines)
